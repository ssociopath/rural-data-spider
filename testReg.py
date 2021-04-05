import time
import requests
import json
import copy
from openpyxl import load_workbook
from reg_get import read_reg_list_txt


def get_html(url, params, headers):
    while True:
        try:
            time.sleep(1)
            req = requests.get(url, params=params, headers=headers, verify=False)
            req.encoding = req.apparent_encoding
            trans_dict = json.loads(req.text)
            break
        except:
            time.sleep(30)
            print('服务器繁忙')
    return trans_dict


def get_time():
    return int(round(time.time() * 1000))


def get_title(project):
    key_value['dfwds'] = '[{{"wdcode":"zb","valuecode":"{}"}}]'.format(project)
    key_value['wds'] = '[{"wdcode":"reg","valuecode":"110000"}]'
    data = get_html(url, headers=headers, params=key_value)
    # 设置表头横坐标
    row = 1
    column = 3
    x_list = data['returndata']['wdnodes'][0]['nodes']
    for i in range(0, len(x_list)):
        print(str(x_list[i]['cname']))
        xls.cell(row=row, column=column + i, value=str(x_list[i]['cname']))


def get_data(reg, project, year):
    # 设置dfwds参数
    key_value['wds'] = '[{{"wdcode":"reg","valuecode":"{}"}}]'.format(reg)
    key_value['dfwds'] = '[{{"wdcode":"zb","valuecode":"{}"}},{{"wdcode":"sj","valuecode":"{}"}}]'.format(project, year)
    # 发起请求
    data = get_html(url, headers=headers, params=key_value)
    # 读取行和列
    row = 1
    column = 1
    for rows in xls:
        if not all([cell.value == None for cell in rows]):
            row += 1
    # 设置表头纵坐标
    xls.cell(row=row, column=column, value=reg_list[key])
    xls.cell(row=row, column=column+1, value=year + '年')
    column = column + 2
    print(reg_list[key] + ' ' + year + '年 ', end='')
    # 读取数据
    rows = []
    data_list = []
    for value in data['returndata']['datanodes']:
        rows.append(value['data']['data'])
        if len(rows) == 10:
            data_list.append(rows.copy())
            rows.clear()
    for i in range(0, len(data_list)):
        for j in range(0, len(data_list[i])):
            xls.cell(row=i + row, column=j + column, value=str(data_list[i][j]))
    print(data_list)


url = 'https://data.stats.gov.cn/easyquery.htm?cn=E0103'
key_value = {}
headers = {'User-Agent': 'Mozilla/5.0(Windows;U;Windows NT6.1;en-US;rv:1.9.1.6) Geko/20091201 Firefox/3.5.6'}
key_value['m'] = 'QueryData'
key_value['dbcode'] = 'fsnd'
key_value['rowcode'] = 'zb'
key_value['colcode'] = 'sj'
key_value['k1'] = str(get_time())

# 读取和写入

requests.packages.urllib3.disable_warnings()
workbook = load_workbook(filename='result/list1.xlsx')
xls = workbook.active
reg_list = read_reg_list_txt()
get_title('A0D05')
for search_year in range(2011, 2020):
    for key in reg_list:
        get_data(key, 'A0D05', str(search_year))
workbook.save(filename='result/list1.xlsx')


