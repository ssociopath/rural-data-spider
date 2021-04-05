import time
import requests
import json
from openpyxl import load_workbook
from reg_get import read_reg_list_txt


def get_html(url, params, headers):
    while True:
        try:
            time.sleep(1)
            req = requests.get(url, params=params, headers=headers, verify=False)
            req.encoding = req.apparent_encoding
            trans_dict = json.loads(req.text)
            break
        except:
            time.sleep(30)
            print('服务器繁忙')
    return trans_dict


def get_time():
    return int(round(time.time() * 1000))


def get_title(project, xls, url, params, headers):
    params['dfwds'] = '[{{"wdcode":"zb","valuecode":"{}"}}]'.format(project)
    params['wds'] = '[{"wdcode":"reg","valuecode":"110000"}]'
    data = get_html(url, headers=headers, params=params)
    # 设置表头横坐标
    row = 1
    column = 3
    x_list = data['returndata']['wdnodes'][0]['nodes']
    for i in range(0, len(x_list)):
        print(str(x_list[i]['cname']))
        xls.cell(row=row, column=column + i, value=str(x_list[i]['cname']))


def get_data(reg, project, year, xls, reg_list, url, params, headers):
    # 设置dfwds参数
    params['wds'] = '[{{"wdcode":"reg","valuecode":"{}"}}]'.format(reg)
    params['dfwds'] = '[{{"wdcode":"zb","valuecode":"{}"}},{{"wdcode":"sj","valuecode":"{}"}}]'.format(project, year)
    # 发起请求
    data = get_html(url, headers=headers, params=params)
    # 读取行和列
    row = 1
    column = 1
    for rows in xls:
        if not all([cell.value == None for cell in rows]):
            row += 1
    # 设置表头纵坐标
    xls.cell(row=row, column=column, value=reg_list[reg])
    xls.cell(row=row, column=column+1, value=year + '年')
    column = column + 2
    print(reg_list[reg] + ' ' + year + '年 ', end='')
    # 读取数据
    rows = []
    for value in data['returndata']['datanodes']:
        rows.append(value['data']['data'])
    for i in range(0, len(rows)):
        xls.cell(row=row, column=column+i, value=str(rows[i]))
    print(rows)


def start_get(project, filename, url, params, headers):
    workbook = load_workbook(filename=filename)
    xls = workbook.active
    reg_list = read_reg_list_txt()
    get_title(project, xls, url, params, headers)
    for search_year in range(2011, 2020):
        for key in reg_list:
            get_data(key, project, str(search_year), xls, reg_list, url, params, headers)
    workbook.save(filename=filename)



