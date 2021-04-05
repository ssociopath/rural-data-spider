import time
import requests
import json
import copy
from openpyxl import load_workbook


def get_time():
    return int(round(time.time() * 1000))


def get_data(param, filename):
    # 设置dfwds参数
    key_value['dfwds'] = param
    # 发起请求
    r = requests.get(url, headers=headers, params=key_value, verify=False)
    data = json.loads(r.text)
    # 读取和写入
    workbook = load_workbook(filename=filename)
    xls = workbook.active
    # 读取行和列
    row = 1
    column = 1
    y_list = ['时间', '2020年', '2019年', '2018年', '2017年', '2016年', '2015年', '2014年', '2013年', '2012年', '2011年']
    x_list = data['returndata']['wdnodes'][0]['nodes']
    for j in range(0, len(y_list)):
        xls.cell(row=j + 1, column=1, value=str(y_list[j]))
    column = column + 1
    for i in range(0, len(x_list)):
        xls.cell(row=row, column=column + i, value=str(x_list[i]['cname']))
    row = row + 1
    # 读取数据
    rows = []
    data_list = []
    for value in data['returndata']['datanodes']:
        rows.append(value['data']['data'])
        if len(rows) == 10:
            data_list.append(rows.copy())
            rows.clear()
    for i in range(0, len(data_list)):
        for j in range(0, len(data_list[i])):
            xls.cell(row=j + row, column=i + column, value=str(data_list[i][j]))
    # 保存文件
    workbook.save(filename=filename)


url = 'https://data.stats.gov.cn/easyquery.htm?cn=C01'
key_value = {}
headers = {'User-Agent': 'Mozilla/5.0(Windows;U;Windows NT6.1;en-US;rv:1.9.1.6) Geko/20091201 Firefox/3.5.6'}
key_value['m'] = 'QueryData'
key_value['dbcode'] = 'hgnd'
key_value['rowcode'] = 'zb'
key_value['colcode'] = 'sj'
key_value['wds'] = '[]'
key_value['k1'] = str(get_time())

get_data('[{"wdcode":"zb","valuecode":"A0D04"},{"wdcode":"sj","valuecode":"LAST10"}]', 'list1.xlsx')


